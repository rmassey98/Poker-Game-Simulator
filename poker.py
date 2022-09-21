# Imports
import random
import pandas as pd
from openpyxl import load_workbook

# Create a playing card class - Format "Ace of Spades"
class PlayingCard():    

    def __init__(self, value, suit):
        self.value = value
        self.suit = suit
    # Represent objects as strings
    def __repr__(self):
        if self.value == 1:
            return "Ace of " + self.suit
        elif self.value == 11:
            return "Jack of " + self.suit
        elif self.value == 12:
            return "Queen of " + self.suit
        elif self.value == 13:
            return "King of " + self.suit
        else:
            return str(self.value) + " of " + self.suit

# Create a standard deck of 52 cards of type PlayingCard
class StandardDeck():
    def __init__(self):
    # Initialise an empty list for deck and community cards
        self.deck = []
        self.community_cards = []

    # Fill deck with playing cards
        suits = ["Hearts", "Spades", "Diamonds", "Clubs"]
        values = [i for i in range(1, 14)]

        for suit in suits:
            for value in values:
                self.deck.append( PlayingCard(value, suit))

    # Methods to shuffle deck, deal a card, and deal community cards
    def shuffle(self):
        random.shuffle(self.deck)

    def deal(self, location):
        location.hand.append(self.deck.pop())

    def deal_community_cards(self):
         for i in range(5):
            self.community_cards.append(self.deck.pop())

# Creating a player class
class Player():
    def __init__(self, name):
        self.hand = []
        self.name = name
        self.score = ""

    def __repr__(self):
        return str(self.name)

# Creating a class that assesses what hand the player has
class PokerScorer():
    def __init__(self, cards):
        self.cards = cards

    def RoyalFlush(self):
        # First determine if there is a flush in the hand
        suits = [card.suit for card in self.cards]
        
        # Determine if ace-high straight can be made with the values that make up the flush
        for suit in suits:
            if suits.count(suit) >= 5:
                values = set([card.value for card in self.cards if card.suit == suit])
                if {1, 10, 11, 12, 13}.issubset(values):
                    return True          

    def StraightFlush(self):
        # First determine if there is a flush in the hand
        suits = [card.suit for card in self.cards]

        # Determine if a straight (excluding ace-high straight) can be made with the values that make up the flush
        for suit in suits:
            if suits.count(suit) >= 5:
                values = set([card.value for card in self.cards if card.suit == suit])

                straight_sets = [{1,2,3,4,5}, {2,3,4,5,6}, {3,4,5,6,7}, {4,5,6,7,8}, {5,6,7,8,9}, {6,7,8,9,10}, {7,8,9,10,11}, {8,9,10,11,12}, {9,10,11,12,13}]
                for straight in straight_sets:
                    if straight.issubset(values):
                        return True

    def FourKind(self):
        values = [card.value for card in self.cards]
        # Check for four of the same value card in the hand
        for value in values:
            if values.count(value) == 4:
                return True

    def FullHouse(self):
        two = False
        three = False
        values = [card.value for card in self.cards]

        # Check for two of the same value card in the hand
        for value in values:
            if values.count(value) == 2:
                two = True
        # Check for three of the same value card in the hand
        for value in values:
            if values.count(value) == 3:
                three = True
        # If both checks satisfied then return True
        if two and three:
            return True


    def Flush(self):
        suits = [card.suit for card in self.cards]

        # Check for 5 of the same suit in the hand
        for suit in suits:
            if suits.count(suit) >= 5:
                return True

    def Straight(self):
        values = [card.value for card in self.cards]
        values_set = set(values)

        # Return false if length of set not long enough
        if not len(values_set) >= 5:
            return False
        # Loop through potential sets and see if it is subset of values
        else:
            straight_sets = [{10,11,12,13,1}, {1,2,3,4,5}, {2,3,4,5,6}, {3,4,5,6,7}, {4,5,6,7,8}, {5,6,7,8,9}, {6,7,8,9,10}, {7,8,9,10,11}, {8,9,10,11,12}, {9,10,11,12,13}]

            for straight in straight_sets:
                if straight.issubset(values_set):
                    return True

    def ThreeKind(self):
        values = [card.value for card in self.cards]
        # Check for three of the same value card in the hand
        for value in values:
            if values.count(value) == 3:
                return True

    def TwoPair(self):
        pairs = []
        values = [card.value for card in self.cards]

        # Checking for different pairs in the hand cards and adding to list of pairs
        for value in values:
            if values.count(value) == 2 and value not in pairs:
                pairs.append(value)

        # Checking if there are at least two pairs in the hand
        if len(pairs) >= 2:
            return True

    def Pair(self):
        values = [card.value for card in self.cards]
        # Checking if there are two of the same value card in hand + community cards
        for value in values:
            if values.count(value) == 2:
                return True

    # High card logic not required as if none of the above checks are true, the only option left if High card
    
    #def HighCard(self):
        #return True


# Creating class to generate poker games spreadsheet
class PlayPoker():
    def __init__(self, num_players, num_iterations, file_path):
        self.num_players = num_players
        self.num_iterations = num_iterations
        self.players = [Player("Player " + str(i+1)) for i in range(self.num_players)]
        self.games = []
        self.file_path = file_path

    def play(self):
        for i in range(self.num_iterations):
            # Start by shuffling the deck
            self.deck = StandardDeck()
            self.deck.shuffle()     

            # Dealing two cards to all players
            for player in self.players:
                self.deck.deal(player)
            for player in self.players:
                self.deck.deal(player)
            # Dealing community cards
            self.deck.deal_community_cards()

            # Checking the score of all players
            for player in self.players:
                combined_hand = player.hand + self.deck.community_cards
                check = PokerScorer(combined_hand)
                if check.RoyalFlush():
                    player.score = "Royal Flush"
                elif check.StraightFlush():
                    player.score = "Straight Flush"
                elif check.FourKind():
                    player.score = "Four of a Kind"
                elif check.FullHouse():
                    player.score = "Full House"
                elif check.Flush():
                    player.score = "Flush"
                elif check.Straight():
                    player.score = "Straight"
                elif check.ThreeKind():
                    player.score = "Three of a Kind"
                elif check.TwoPair():
                    player.score = "Two Pair"
                elif check.Pair():
                    player.score = "Pair"
                else:
                    player.score = "High Card"

            # Creating output format
            headings = ["Card 1", "Card 2", "Community Card 1", "Community Card 2", "Community Card 3", "Community Card 4", "Community Card 5", "Score"]
            self.games.append(headings)
    
            for player in self.players:
                self.games.append(player.hand + self.deck.community_cards + [player.score])
                player.hand = []

        # Outputting to excel document
        Poker_df = pd.DataFrame(self.games)
        Poker_df.to_excel(self.file_path,index=False, header=False)  

        # Formatting spreadsheet with countifs and percentages of each hands
        wb = load_workbook(self.file_path)
        ws = wb.active

        ws['J2'] = "Hand"
        ws['J3'] = "High Card"
        ws['J4'] = "Pair"
        ws['J5'] = "Two Pair"
        ws['J6'] = "Three of a Kind"
        ws['J7'] = "Straight"
        ws['J8'] = "Flush"
        ws['J9'] = "Full House"
        ws['J10'] = "Four of a Kind"
        ws['J11'] = "Straight Flush"
        ws['J12'] = "Royal Flush"
        ws['J14'] = "Games"
        ws['J15'] = "Players"
        ws['J16'] = "Hands"

        ws['K2'] = "Count"
        ws['K3'] = "=COUNTIF(H:H,J3)"
        ws['K4'] = "=COUNTIF(H:H,J4)"
        ws['K5'] = "=COUNTIF(H:H,J5)"
        ws['K6'] = "=COUNTIF(H:H,J6)"
        ws['K7'] = "=COUNTIF(H:H,J7)"
        ws['K8'] = "=COUNTIF(H:H,J8)"
        ws['K9'] = "=COUNTIF(H:H,J9)"
        ws['K10'] = "=COUNTIF(H:H,J10)"
        ws['K11'] = "=COUNTIF(H:H,J11)"
        ws['K12'] = "=COUNTIF(H:H,J12)"
        ws['K14'] = "=COUNTIF(H:H,$H$1)"
        ws['K15'] = str(self.num_players)
        ws['K16'] = "=K14*K15"

        ws['L2'] = "Percentage"
        ws['L3'] = "=K3/$K$16"
        ws['L4'] = "=K4/$K$16"
        ws['L5'] = "=K5/$K$16"
        ws['L6'] = "=K6/$K$16"
        ws['L7'] = "=K7/$K$16"
        ws['L8'] ="=K8/$K$16"
        ws['L9'] = "=K9/$K$16"
        ws['L10'] = "=K10/$K$16"
        ws['L11'] = "=K11/$K$16"
        ws['L12'] = "=K12/$K$16"

        ws['L3'].number_format = '0.0000%'
        ws['L4'].number_format = '0.0000%'
        ws['L5'].number_format = '0.0000%'
        ws['L6'].number_format = '0.0000%'
        ws['L7'].number_format = '0.0000%'
        ws['L8'].number_format = '0.0000%'
        ws['L9'].number_format = '0.0000%'
        ws['L10'].number_format = '0.0000%'
        ws['L11'].number_format = '0.0000%'
        ws['L12'].number_format = '0.0000%'

        wb.save(self.file_path)

        print("Successfully created excel document containing results of {}, {} player games of poker.".format(self.num_iterations, self.num_players))
















        
        












